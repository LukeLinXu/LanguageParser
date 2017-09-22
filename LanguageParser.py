# -*- coding: utf-8 -*-
import os
from xml.dom import minidom
from xml.etree.ElementTree import Element, SubElement, tostring, ElementTree
from openpyxl import load_workbook
import sys

__author__ = 'llin'


def check_duplicate_id(lists, platform):
    for item in lists:
        hash_set = set()
        for value in item:
            id = value[0]
            if id in hash_set:
                print("check id", platform, "fail for duplicate id", id)
                break
            else:
                hash_set.add(id)
        else:
            print("check id", platform, "pass")


def read_excel(file='LanguageParser.xlsx'):
    wb = load_workbook(filename = file)
    # pick the first sheet
    sheet_ranges = wb[wb._sheets[0].title]

    column_position = 1
    lang_list = 0
    while True:
        title = sheet_ranges.cell(row = 1, column = column_position)
        if title.value == None: break
        if lang_list != 0:
            lang_list.append(title.value.lower())
        if title.value == 'Id':
            lang_list = list()
        column_position = column_position + 1

    for item in lang_list:
        print(item)

    android_lists = [list() for item in lang_list]
    ios_lists = [list() for item in lang_list]
    windows_lists = [list() for item in lang_list]

    rowcount = 2
    for row in sheet_ranges.rows:
        position = column_position - len(lang_list) - 2
        platform = sheet_ranges.cell(row = rowcount, column = position)
        id = sheet_ranges.cell(row = rowcount, column = position+1)
        position = position + 2
        cells = []
        for item in lang_list:
            cell = sheet_ranges.cell(row = rowcount, column = position)
            value = remove_hyper_link(cell.value)
            if value != None:
                value = value.replace('\\"', '\"')
            cells.append(value)
            position = position + 1
        if id.value == None:
            break
        rowcount = rowcount + 1
        if  platform.value != None:
            if 'COMMON' in platform.value:
                for map, cell in zip(android_lists, cells):
                    if cell != None:
                        map.append((id.value, android_formatter(cell)))
                for map, cell in zip(ios_lists, cells):
                    if cell != None:
                        map.append((id.value, ios_formatter(cell)))
                for map, cell in zip(windows_lists, cells):
                    if cell != None:
                        map.append((id.value, windows_formatter(cell)))
            else:
                if "ANDROID" in platform.value:
                    for map, cell in zip(android_lists, cells):
                        if cell != None:
                            map.append((id.value, android_formatter(cell)))

                if "IOS" in platform.value:
                    for map, cell in zip(ios_lists, cells):
                        if cell != None :
                            map.append((id.value, ios_formatter(cell)))

                if "WINDOWS" in platform.value:
                    for map, cell in zip(windows_lists, cells):
                        if cell != None:
                            map.append((id.value, windows_formatter(cell)))

    check_duplicate_id(android_lists, "android")
    check_duplicate_id(ios_lists, "ios")
    check_duplicate_id(windows_lists, "windows")

    for map, lang in zip(android_lists, lang_list):
        # map = sorted(map.items())
        createAndroidFile(map, lang)
    for map, lang in zip(ios_lists, lang_list):
        # map = sorted(map.items())
        createiOSFile(map, lang)
    for map, lang in zip(windows_lists, lang_list):
        # map = sorted(map.items())
        createWindowsFile(map, lang)

def prettify(elem):
    rough_string = tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="    ", encoding='utf-8')


def createAndroidFile(map, initial):
    top = Element('resources')
    for key, value in map:
        value = android_escape(value)
        child = SubElement(top, 'string', {'name':key})
        child.text = value
    foldername = 'values'
    if 'zh' in initial:
        foldername = foldername+'-zh'
    elif 'fr' in initial:
        foldername = foldername+'-fr'
    path = 'Android'+os.sep+foldername
    os.makedirs(path, exist_ok=True)
    with open(path+os.sep+'strings.xml', 'w', encoding="utf-8") as f:
        f.write(prettify(top).decode('utf-8'))
        f.close()

def createWindowsFile(map, initial):
    top = Element('MessageStore')
    if 'zh' in initial:
        initial = 'zh-CN'
        top.set('EnglishName', 'Chinese')
        top.set('CultureName', 'Chinese')
    elif 'fr' in initial:
        initial = 'fr-FR'
        top.set('EnglishName', 'French')
        top.set('CultureName', 'French')
    else:
        initial = 'en-US'
        top.set('EnglishName', 'English')
        top.set('CultureName', 'English')
    for key, value in map:
        child = SubElement(top, 'Message', {'id':key})
        child.text = value
    filename = 'MessageStore.'
    filename = filename + initial
    path = 'Windows'
    os.makedirs(path, exist_ok=True)
    with open(path+os.sep+filename+'.xml', 'w', encoding="utf-8") as f:
        f.write(prettify(top).decode('utf-8'))
        f.close()

def createiOSFile(map, initial):
    foldername = '.lproj'
    if 'zh' in initial:
        initial = 'zh'
    elif 'fr' in initial:
        initial = 'fr'
    else:
        initial = 'en'
    foldername = initial+foldername

    # create for ObjectC
    path = 'iOS'+os.sep+'ObjectC'+os.sep+foldername
    os.makedirs(path, exist_ok=True)
    content = []
    for key, value in map:
        value = ios_escape(value)
        content.append('''"{0}" = "{1}";\n'''.format(key, value))
    content = ''.join(content)
    with open(path+os.sep+'Localizable.strings', 'w', encoding="utf-8") as f:
        f.write(content)
        f.close()

    # create for swift
    path = 'iOS'+os.sep+'Swift'+os.sep+foldername
    os.makedirs(path, exist_ok=True)
    content = []
    for key, value in map:
        value = ios_escape(value)
        content.append('''{0} = "{1}"\n'''.format(key, value))
    content = ''.join(content)
    with open(path+os.sep+'Localizable.strings', 'w', encoding="utf-8") as f:
        f.write(content)
        f.close()

def remove_hyper_link(data):
    if data != None and '=HYPERLINK' in data:
        return data.split('"')[-2]
    else:
        return data

def android_formatter(data):
    position = 1
    while '%S%' in data or '%D%' in data or '%L%' in data:
        pos_s = data.find('%S%')
        pos_d = data.find('%D%')
        pos_l = data.find('%L%')
        pos = min([x for x in [pos_d, pos_s, pos_l] if x != -1])
        if pos_s == pos:
            data = data.replace('%S%', '%'+str(position)+'$s', 1)
        elif pos_d == pos:
            data = data.replace('%D%', '%' + str(position) + '$d', 1)
        elif pos_l == pos:
            data = data.replace('%L%', '%' + str(position) + '$l', 1)
        position += 1
    return data

def windows_formatter(data):
    data = data.replace('%L%', '%S%')
    data = data.replace('%D%', '%S%')
    position = 0
    while '%S%' in data:
        data = data.replace('%S%', '{'+str(position)+'}', 1)
        position = position+1
    return data

def ios_escape(text):
    if '''"''' in text:
        text = text.replace('''"''', '''\\"''')
    return text

def android_escape(text):
    if """'""" in text:
        text = text.replace("""'""", """\\'""")
    return text

def ios_formatter(data):
    data = data.replace('%L%', '%lu')
    data = data.replace('%S%', '%@')
    data = data.replace('%D%', '%d')
    return data

if __name__ == '__main__':
    if(len(sys.argv)<2):
        read_excel()
    else:
        read_excel(sys.argv[1])
